from flask import Flask , render_template, request
import pandas as pd
from cryptography.fernet import Fernet
# import encryption

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

app = Flask(__name__)
key = b'7qt6LTeQ5L9dFO4KUTvwFE9Px2b5tRVg0wIXRqSUvL0='
fernet = Fernet(key)

# df = pd.read_excel('Database.xlsx')

@app.route("/",methods=['GET', 'POST'])
def login():

    try:
        workbook_name = 'Database.xlsx'
        wb = load_workbook(workbook_name)
        page = wb.active
    except Exception:
        headers       = ['Email','Password','Address','City','state','Zip']
        workbook_name = 'Database.xlsx'
        wb = Workbook()
        page = wb.active
        page.title = 'Database'
        page.append(headers) # write the headers to the first line

        wb.save(filename = workbook_name)


    

    # try:
    # df = pd.read_excel('DataBase.xlsx')
    # print("Database already exists")
    # data = "Database already exists"
    # except Exception:
    #     df = pd.DataFrame({'Email':[],
    #                     'Password':[],
    #                     'Address':[],
    #                     'City':[],
    #                     'State':[],
    #                     'Zip':[]})
    #     writer = pd.ExcelWriter('DataBase.xlsx', engine='xlsxwriter')

    #     # Convert the dataframe to an XlsxWriter Excel object.
    #     df.to_excel(writer, sheet_name='Sheet1', index=False)
    #     writer.save()

        # df.to_csv('DataBase.csv',index=False)
        # data ="Database created sucessfully!!!" 
        # print("Database created sucessfully!!!")

    return render_template('login.html') 

@app.route("/register",methods=['GET', 'POST'])
def register():
        # print(df)
    df = pd.read_excel('Database.xlsx')

    return render_template('register.html',df=df) 



@app.route("/logcheck",methods=['GET', 'POST'])
def logcheck():
    df = pd.read_excel('Database.xlsx')
    a = 0
    no = 0
    index = 0
    lst = ['Email','Password','Address','City','state','Zip']
    dict1 = {}
    decr = {}
    if request.method=='POST':
        data = request.form
        for i in df['Email']:
            j = bytes(i, 'utf-8')
            if request.form['Email'] == fernet.decrypt(j).decode():
                a=1
                index = no
                # print("index : ",index)
            else:
                no = no + 1
        if request.form['Password'] == fernet.decrypt(bytes(df['Password'][index], 'utf-8')).decode():
            a = 2
        for i in lst:
            dict1[i] = df[i][index]
            decr[i] = fernet.decrypt(bytes(df[i][index], 'utf-8')).decode()
        
        # for i in df:
        #     if request.form['Email'] == fernet.decrypt(bytes(df[i], 'utf-8')).decode():
        if len(request.form['Email']) == 0:
            a = 0      

    if a == 2:
        # return "{{index}}"
        return render_template('home.html',dict1=dict1,decr=decr)
    else:
        return render_template('notfound.html')

@app.route("/home",methods=['GET', 'POST'])
def home():
    df = pd.read_excel('Database.xlsx')
    if request.method=='POST':

        a = 0
        
        for i in df['Email']:
            if request.form['Email'] == fernet.decrypt(bytes(i, 'utf-8')).decode():
                a = 1
            
        
        # encMessage = 'gAAAAABiABrw9mEsHNB6GCMNF6r7_mX6Og6EOkTLfd4NfLHaYwGiv5vAazHOyHvz-7cVQwdvD9hcZC-lgUPDWsMwEftdLqeTwrd_tVShDORkZrPX7awckk4='
        # decMessage = fernet.decrypt(encMessage).decode()
 
        # print("decrypted string: ", decMessage)
        # encMessage = fernet.encrypt(message.encode())
        if a != 1:
            data = []
            dict1 = {}
            decr = {}
            for i in request.form:
                j = fernet.encrypt(request.form[i].encode())
                data.append(j) 
                dict1[i]=j

                decr[i]=fernet.decrypt(j).decode()
            # print("original : ",request.form[i])
            # print("=================================================================================")
            # print(j)
            # print("=================================================================================")
            # print("decr : ",fernet.decrypt(j).decode())
            # print("=================================================================================")
            
        # print(data)

            workbook_name = 'Database.xlsx'
            wb = load_workbook(workbook_name)
            page = wb.active

        # New data to write:
        # for i in data:
        #     print(i)
            page.append(data)

            wb.save(filename=workbook_name) 

        # for i in range(0,7):
        #     for j in df:
        #         k = df[j][i]
        #         enc = bytes(k, 'utf-8')
        #         decMessage = fernet.decrypt(enc).decode()
        #         print(k)
        #         print("==========================================================")
        #         print(decMessage)
        #         print("==========================================================")
            return render_template('home.html',dict1=dict1,decr=decr) 
        else:
            return render_template('exist.html')

if __name__ == "__main__":
    app.run(debug = True)